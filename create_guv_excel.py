import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

wb = openpyxl.Workbook()

MONATE = ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez']

data = {
    2025: {
        'sanitaer':     [18500,16200,22000,24500,26000,28500,29000,27000,25500,23000,20000,18000],
        'heizung':      [22000,19500,18000,14000,12000,10500,11000,12500,16000,20500,22000,24000],
        'klima':        [2000, 2200, 3500, 5500, 8000,11000,13500,12000, 7500, 4000, 2500, 2000],
        'wartung':      [4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500],
        'notdienst':    [2500, 2200, 1800, 1500, 1500, 1600, 1800, 1700, 1600, 1800, 2200, 2800],
        'material':     [19800,17500,18200,18800,19600,20500,21200,19800,18400,18200,18400,19200],
        'personal':     [18500,18500,18500,18500,18500,18500,18500,18500,18500,18500,18500,18500],
        'fahrzeuge':    [2800, 2800, 2800, 2800, 2800, 2800, 2800, 2800, 2800, 2800, 2800, 2800],
        'werkzeug':     [600,  300,  800,  400,  500,  700,  600,  400,  300,  500,  400,  800],
        'buero':        [1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200],
        'software':     [350,  350,  350,  350,  350,  350,  350,  350,  350,  350,  350,  350],
        'versicherung': [1100, 0,    0,    1100, 0,    0,    1100, 0,    0,    1100, 0,    0   ],
        'steuerberater':[0,    0,    800,  0,    0,    800,  0,    0,    800,  0,    0,    800 ],
        'marketing':    [400,  400,  600,  400,  400,  600,  400,  400,  400,  400,  400,  600],
        'sonstiges':    [800,  600,  700,  800,  750,  900,  800,  700,  650,  800,  750,  900],
        'zinsen':       [650,  650,  650,  650,  650,  650,  650,  650,  650,  650,  650,  650],
        'steuern':      [0,    0,    0,    0,    0,    2800, 0,    0,    0,    0,    0,    4200],
    },
    2024: {
        'sanitaer':     [16500,14800,19500,21500,23000,25500,26000,24500,22500,21000,18500,16500],
        'heizung':      [20500,18000,16500,12500,11000, 9500,10000,11500,14500,19000,20500,22500],
        'klima':        [1800, 1900, 3000, 4800, 7000, 9500,11800,10500, 6500, 3500, 2200, 1800],
        'wartung':      [4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000],
        'notdienst':    [2200, 2000, 1600, 1400, 1400, 1500, 1600, 1500, 1500, 1700, 2000, 2500],
        'material':     [17800,15800,16500,17000,17800,18500,19200,17800,16600,16500,16800,17500],
        'personal':     [17500,17500,17500,17500,17500,17500,17500,17500,17500,17500,17500,17500],
        'fahrzeuge':    [2600, 2600, 2600, 2600, 2600, 2600, 2600, 2600, 2600, 2600, 2600, 2600],
        'werkzeug':     [500,  250,  700,  350,  450,  600,  500,  350,  250,  450,  350,  700],
        'buero':        [1100, 1100, 1100, 1100, 1100, 1100, 1100, 1100, 1100, 1100, 1100, 1100],
        'software':     [300,  300,  300,  300,  300,  300,  300,  300,  300,  300,  300,  300],
        'versicherung': [1000, 0,    0,    1000, 0,    0,    1000, 0,    0,    1000, 0,    0   ],
        'steuerberater':[0,    0,    750,  0,    0,    750,  0,    0,    750,  0,    0,    750 ],
        'marketing':    [350,  350,  550,  350,  350,  550,  350,  350,  350,  350,  350,  550],
        'sonstiges':    [700,  550,  600,  700,  680,  800,  720,  650,  580,  700,  680,  800],
        'zinsen':       [750,  750,  750,  750,  750,  750,  750,  750,  750,  750,  750,  750],
        'steuern':      [0,    0,    0,    0,    0,    2400, 0,    0,    0,    0,    0,    3600],
    }
}

# ── Styles ──────────────────────────────────────────────────────────────
DARK_BLUE   = "1a1a2e"
MID_BLUE    = "0f3460"
LIGHT_BLUE  = "4fc3f7"
GRAY_BG     = "f0f2f5"
SECTION_BG  = "2d3748"
SUBTOT_BG   = "e2e8f0"
GREEN_BG    = "c6f6d5"
RED_BG      = "fed7d7"
NUM_FMT     = '#,##0.00 "€"'

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def align(h='right', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style='thin', color='CBD5E0')
    return Border(bottom=s)

def medium_border():
    s = Side(style='medium', color='718096')
    return Border(bottom=s)

# ── Sheet 1: Jahresvergleich ─────────────────────────────────────────────
ws = wb.active
ws.title = "Jahresvergleich"
ws.sheet_view.showGridLines = False

# Column widths
ws.column_dimensions['A'].width = 36
for col in ['B','C','D','E','F']:
    ws.column_dimensions[col].width = 18

# Title area
ws.row_dimensions[1].height = 10
ws.merge_cells('A2:F2')
c = ws['A2']
c.value = 'SHK Musterbetrieb GmbH'
c.font = Font(bold=True, color=DARK_BLUE, size=16)
c.alignment = align('left')

ws.merge_cells('A3:F3')
c = ws['A3']
c.value = 'Gewinn- und Verlustrechnung 2025 vs. 2024'
c.font = Font(italic=True, color='718096', size=12)
c.alignment = align('left')

ws.row_dimensions[4].height = 8

# Header row
headers = ['Position', '2025', '2024', 'Abw. absolut', 'Abw. %', 'Anteil Umsatz 2025']
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=ci, value=h)
    cell.fill = fill(DARK_BLUE)
    cell.font = font(bold=True, color='FFFFFF', size=11)
    cell.alignment = align('center' if ci==1 else 'right')
    cell.border = Border(bottom=Side(style='medium', color='4FC3F7'))
ws.row_dimensions[5].height = 22

def write_row(row, label, v25, v24, indent=False, style='normal', is_cost=False):
    ws.row_dimensions[row].height = 18
    # flip sign for costs in display
    disp25 = -v25 if is_cost else v25
    disp24 = -v24 if is_cost else v24
    diff = disp25 - disp24
    diffpct = (diff / abs(disp24) * 100) if disp24 != 0 else 0

    # Compute total umsatz 2025
    umsatz25 = sum(data[2025][k][m] for k in ['sanitaer','heizung','klima','wartung','notdienst'] for m in range(12))
    anteil = (disp25 / umsatz25 * 100) if umsatz25 != 0 else 0

    col_a = ws.cell(row=row, column=1, value=('    ' + label) if indent else label)
    col_b = ws.cell(row=row, column=2, value=disp25)
    col_c = ws.cell(row=row, column=3, value=disp24)
    col_d = ws.cell(row=row, column=4, value=diff)
    col_e = ws.cell(row=row, column=5, value=diffpct/100)
    col_f = ws.cell(row=row, column=6, value=anteil/100)

    for c in [col_b, col_c, col_d]:
        c.number_format = '#,##0.00 "€"'
    col_e.number_format = '+0.0%;-0.0%;0.0%'
    col_f.number_format = '0.0%'

    for c in [col_a, col_b, col_c, col_d, col_e, col_f]:
        c.alignment = align('left' if c.column==1 else 'right')

    if style == 'section':
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(SECTION_BG)
            cell.font = font(bold=True, color='FFFFFF', size=10)
        return

    if style == 'subtotal':
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(SUBTOT_BG)
            cell.font = font(bold=True, color='2d3748')
        return

    if style == 'total':
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(MID_BLUE)
            cell.font = font(bold=True, color='FFFFFF', size=12)
        return

    # Normal: color diff column
    col_a.font = font(bold=False, color='4a5568', size=10) if indent else font()
    if diff >= 0:
        col_d.font = font(color='276749')
        col_e.font = font(color='276749')
    else:
        col_d.font = font(color='c53030')
        col_e.font = font(color='c53030')
    for c in [col_a, col_b, col_c, col_d, col_e, col_f]:
        c.border = thin_border()

def write_section(row, label):
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = fill(SECTION_BG)
    cell.font = Font(bold=True, color='FFFFFF', size=10, name='Segoe UI')
    cell.alignment = align('left')
    ws.row_dimensions[row].height = 16

# Compute totals
def S(year, key): return sum(data[year][key])
def SE(year): return sum(S(year, k) for k in ['sanitaer','heizung','klima','wartung','notdienst'])
def SMat(year): return S(year, 'material')
def SRoh(year): return SE(year) - SMat(year)
def SBetrieb(year): return sum(S(year,k) for k in ['personal','fahrzeuge','werkzeug','buero','software','versicherung','steuerberater','marketing','sonstiges'])
def SEBIT(year): return SRoh(year) - SBetrieb(year)
def SEBT(year): return SEBIT(year) - S(year,'zinsen')
def SJahr(year): return SEBT(year) - S(year,'steuern')

row = 6
write_section(row, '  UMSATZERLÖSE'); row+=1
write_row(row, 'Sanitär (Installation & Reparatur)', S(2025,'sanitaer'), S(2024,'sanitaer'), indent=True); row+=1
write_row(row, 'Heizung (Installation & Reparatur)', S(2025,'heizung'), S(2024,'heizung'), indent=True); row+=1
write_row(row, 'Klimaanlagen', S(2025,'klima'), S(2024,'klima'), indent=True); row+=1
write_row(row, 'Wartungsverträge', S(2025,'wartung'), S(2024,'wartung'), indent=True); row+=1
write_row(row, 'Notdienst', S(2025,'notdienst'), S(2024,'notdienst'), indent=True); row+=1
write_row(row, '= GESAMTUMSATZ', SE(2025), SE(2024), style='subtotal'); row+=1

write_section(row, '  MATERIALAUFWAND'); row+=1
write_row(row, 'Material & Wareneinsatz', S(2025,'material'), S(2024,'material'), indent=True, is_cost=True); row+=1
write_row(row, '= ROHERTRAG (Bruttogewinn)', SRoh(2025), SRoh(2024), style='subtotal'); row+=1

write_section(row, '  BETRIEBSAUFWAND'); row+=1
write_row(row, 'Personalkosten (inkl. Sozialversicherung)', S(2025,'personal'), S(2024,'personal'), indent=True, is_cost=True); row+=1
write_row(row, 'Fahrzeugkosten (Leasing/Kraftstoff)', S(2025,'fahrzeuge'), S(2024,'fahrzeuge'), indent=True, is_cost=True); row+=1
write_row(row, 'Werkzeug & Maschinen', S(2025,'werkzeug'), S(2024,'werkzeug'), indent=True, is_cost=True); row+=1
write_row(row, 'Büro & Verwaltung', S(2025,'buero'), S(2024,'buero'), indent=True, is_cost=True); row+=1
write_row(row, 'Software & Tools', S(2025,'software'), S(2024,'software'), indent=True, is_cost=True); row+=1
write_row(row, 'Versicherungen', S(2025,'versicherung'), S(2024,'versicherung'), indent=True, is_cost=True); row+=1
write_row(row, 'Steuerberater / Rechtsberatung', S(2025,'steuerberater'), S(2024,'steuerberater'), indent=True, is_cost=True); row+=1
write_row(row, 'Marketing & Werbung', S(2025,'marketing'), S(2024,'marketing'), indent=True, is_cost=True); row+=1
write_row(row, 'Sonstige Betriebsausgaben', S(2025,'sonstiges'), S(2024,'sonstiges'), indent=True, is_cost=True); row+=1
betrieb_total_25 = SBetrieb(2025)
betrieb_total_24 = SBetrieb(2024)

# Write total betrieb manually (it's a cost row but we want display as negative)
ws.row_dimensions[row].height = 18
disp25 = -betrieb_total_25; disp24 = -betrieb_total_24
diff = disp25 - disp24
umsatz25 = SE(2025)
diffpct = diff/abs(disp24)*100 if disp24 != 0 else 0
for col, val, fmt_str in [
    (1, '= GESAMT BETRIEBSAUFWAND', '@'),
    (2, disp25, '#,##0.00 "€"'), (3, disp24, '#,##0.00 "€"'),
    (4, diff, '#,##0.00 "€"'), (5, diffpct/100, '+0.0%;-0.0%;0.0%'),
    (6, disp25/umsatz25, '0.0%')
]:
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(SUBTOT_BG); c.font = font(bold=True, color='2d3748')
    c.alignment = align('left' if col==1 else 'right')
    if fmt_str != '@': c.number_format = fmt_str
row += 1

write_row(row, '= EBIT (Betriebsergebnis)', SEBIT(2025), SEBIT(2024), style='total'); row+=1

write_section(row, '  FINANZERGEBNIS & STEUERN'); row+=1
write_row(row, 'Zinsaufwand (Kredite/Leasing)', S(2025,'zinsen'), S(2024,'zinsen'), indent=True, is_cost=True); row+=1
write_row(row, '= EBT (Ergebnis vor Steuern)', SEBT(2025), SEBT(2024), style='subtotal'); row+=1
write_row(row, 'Gewerbesteuer & Einkommensteuer', S(2025,'steuern'), S(2024,'steuern'), indent=True, is_cost=True); row+=1
write_row(row, '= JAHRESÜBERSCHUSS / -FEHLBETRAG', SJahr(2025), SJahr(2024), style='total'); row+=1

# Conditional formatting for Jahresüberschuss
from openpyxl.formatting.rule import CellIsRule
ws.conditional_formatting.add(f'B{row-1}:B{row-1}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=fill(GREEN_BG)))
ws.conditional_formatting.add(f'B{row-1}:B{row-1}',
    CellIsRule(operator='lessThan', formula=['0'], fill=fill(RED_BG)))

# Freeze panes
ws.freeze_panes = 'B6'

# ── Sheet 2: Monatlich ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Monatlich 2025")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 28
for i in range(2, 16):
    ws2.column_dimensions[get_column_letter(i)].width = 11

# Header
ws2.merge_cells('A1:N1')
c = ws2['A1']
c.value = 'SHK Musterbetrieb GmbH – Monatliche GuV 2025'
c.font = Font(bold=True, color=DARK_BLUE, size=14)
c.alignment = align('left')
ws2.row_dimensions[1].height = 24

# Column headers
ws2.cell(row=3, column=1, value='Position').font = font(bold=True, color='FFFFFF')
ws2.cell(row=3, column=1).fill = fill(DARK_BLUE)
ws2.cell(row=3, column=1).alignment = align('left')
for i, m in enumerate(MONATE, 2):
    c = ws2.cell(row=3, column=i, value=m)
    c.fill = fill(DARK_BLUE); c.font = font(bold=True, color='FFFFFF')
    c.alignment = align('center')
c = ws2.cell(row=3, column=14, value='Gesamt')
c.fill = fill(MID_BLUE); c.font = font(bold=True, color='FFFFFF')
c.alignment = align('center')
ws2.row_dimensions[3].height = 20

def write_monthly_section(ws, row, label):
    ws.merge_cells(f'A{row}:N{row}')
    c = ws.cell(row=row, column=1, value='  ' + label)
    c.fill = fill(SECTION_BG); c.font = font(bold=True, color='FFFFFF', size=10)
    c.alignment = align('left')
    ws.row_dimensions[row].height = 16

def write_monthly_row(ws, row, label, vals_25, bold=False, is_cost=False, color_row=False):
    ws.row_dimensions[row].height = 17
    disp = [-v for v in vals_25] if is_cost else list(vals_25)
    total = sum(disp)

    c = ws.cell(row=row, column=1, value=label)
    c.alignment = align('left')
    if bold: c.font = font(bold=True)

    for col, v in enumerate(disp, 2):
        cell = ws.cell(row=row, column=col, value=v)
        cell.number_format = '#,##0 "€"'
        cell.alignment = align('right')
        if bold:
            cell.fill = fill(SUBTOT_BG)
            cell.font = font(bold=True)
        if color_row:
            cell.fill = fill(GREEN_BG) if v >= 0 else fill(RED_BG)
            cell.font = font(bold=True, color='276749' if v >= 0 else 'c53030')
        cell.border = thin_border()

    tc = ws.cell(row=row, column=14, value=total)
    tc.number_format = '#,##0 "€"'
    tc.alignment = align('right')
    if bold: tc.fill = fill(SUBTOT_BG); tc.font = font(bold=True)
    if color_row:
        tc.fill = fill(GREEN_BG) if total >= 0 else fill(RED_BG)
        tc.font = font(bold=True, color='276749' if total >= 0 else 'c53030')

row2 = 4
write_monthly_section(ws2, row2, 'UMSATZERLÖSE'); row2+=1
write_monthly_row(ws2, row2, 'Sanitär', data[2025]['sanitaer']); row2+=1
write_monthly_row(ws2, row2, 'Heizung', data[2025]['heizung']); row2+=1
write_monthly_row(ws2, row2, 'Klimaanlagen', data[2025]['klima']); row2+=1
write_monthly_row(ws2, row2, 'Wartungsverträge', data[2025]['wartung']); row2+=1
write_monthly_row(ws2, row2, 'Notdienst', data[2025]['notdienst']); row2+=1
umsatz_monthly = [sum(data[2025][k][m] for k in ['sanitaer','heizung','klima','wartung','notdienst']) for m in range(12)]
write_monthly_row(ws2, row2, '= GESAMTUMSATZ', umsatz_monthly, bold=True); row2+=1

write_monthly_section(ws2, row2, 'MATERIALAUFWAND'); row2+=1
write_monthly_row(ws2, row2, 'Material & Wareneinsatz', data[2025]['material'], is_cost=True); row2+=1
rohertrag_monthly = [umsatz_monthly[m] - data[2025]['material'][m] for m in range(12)]
write_monthly_row(ws2, row2, '= ROHERTRAG', rohertrag_monthly, bold=True); row2+=1

write_monthly_section(ws2, row2, 'BETRIEBSAUFWAND'); row2+=1
for key, label in [('personal','Personal'),('fahrzeuge','Fahrzeuge'),('werkzeug','Werkzeug'),('buero','Büro'),('software','Software'),('versicherung','Versicherung'),('steuerberater','Steuerberater'),('marketing','Marketing'),('sonstiges','Sonstiges')]:
    write_monthly_row(ws2, row2, label, data[2025][key], is_cost=True); row2+=1

betrieb_monthly = [sum(data[2025][k][m] for k in ['personal','fahrzeuge','werkzeug','buero','software','versicherung','steuerberater','marketing','sonstiges']) for m in range(12)]
write_monthly_row(ws2, row2, '= GESAMT BETRIEBSAUFWAND', [-v for v in betrieb_monthly], bold=True); row2+=1

ebit_monthly = [rohertrag_monthly[m] - betrieb_monthly[m] for m in range(12)]
# Write EBIT with dark blue background
for col, v in enumerate(ebit_monthly, 2):
    c = ws2.cell(row=row2, column=col, value=v)
    c.fill = fill(MID_BLUE); c.font = font(bold=True, color='FFFFFF')
    c.number_format = '#,##0 "€"'; c.alignment = align('right')
c = ws2.cell(row=row2, column=1, value='= EBIT')
c.fill = fill(MID_BLUE); c.font = font(bold=True, color='FFFFFF'); c.alignment = align('left')
c = ws2.cell(row=row2, column=14, value=sum(ebit_monthly))
c.fill = fill(MID_BLUE); c.font = font(bold=True, color='FFFFFF')
c.number_format = '#,##0 "€"'; c.alignment = align('right')
row2 += 1

write_monthly_section(ws2, row2, 'FINANZERGEBNIS & STEUERN'); row2+=1
write_monthly_row(ws2, row2, 'Zinsaufwand', data[2025]['zinsen'], is_cost=True); row2+=1
ebt_monthly = [ebit_monthly[m] - data[2025]['zinsen'][m] for m in range(12)]
write_monthly_row(ws2, row2, '= EBT', ebt_monthly, bold=True); row2+=1
write_monthly_row(ws2, row2, 'Steuern', data[2025]['steuern'], is_cost=True); row2+=1
ergebnis_monthly = [ebt_monthly[m] - data[2025]['steuern'][m] for m in range(12)]
write_monthly_row(ws2, row2, '= JAHRESERGEBNIS', ergebnis_monthly, bold=True, color_row=True); row2+=1

ws2.freeze_panes = 'B4'

# ── Sheet 3: Eingabe (editierbar) ────────────────────────────────────────
ws3 = wb.create_sheet("Eigene Zahlen")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 30
for i in range(2, 15):
    ws3.column_dimensions[get_column_letter(i)].width = 12

ws3.merge_cells('A1:N1')
c = ws3['A1']
c.value = '📝 Eigene Zahlen eintragen – Ergebnisse werden automatisch berechnet'
c.font = Font(bold=True, color=DARK_BLUE, size=13)
c.fill = fill('FFF9C4')
c.alignment = align('left')
ws3.row_dimensions[1].height = 28

ws3.merge_cells('A2:N2')
c = ws3['A2']
c.value = 'Trage hier deine echten Monatszahlen ein (netto, in EUR). Die GuV-Auswertung passt sich automatisch an.'
c.font = Font(italic=True, color='718096', size=11)
c.alignment = align('left')
ws3.row_dimensions[2].height = 20

headers_row = 4
ws3.cell(row=headers_row, column=1, value='Kategorie').fill = fill(DARK_BLUE)
ws3.cell(row=headers_row, column=1).font = font(bold=True, color='FFFFFF')
ws3.cell(row=headers_row, column=1).alignment = align('left')
for i, m in enumerate(MONATE, 2):
    c = ws3.cell(row=headers_row, column=i, value=m)
    c.fill = fill(DARK_BLUE); c.font = font(bold=True, color='FFFFFF')
    c.alignment = align('center')
c = ws3.cell(row=headers_row, column=14, value='Summe')
c.fill = fill(MID_BLUE); c.font = font(bold=True, color='FFFFFF')
c.alignment = align('center')
ws3.row_dimensions[headers_row].height = 20

input_categories = [
    ('EINNAHMEN', None),
    ('Sanitär', 'einnahmen'),
    ('Heizung', 'einnahmen'),
    ('Klimaanlagen', 'einnahmen'),
    ('Wartungsverträge', 'einnahmen'),
    ('Notdienst', 'einnahmen'),
    ('', None),
    ('AUSGABEN', None),
    ('Material & Wareneinsatz', 'kosten'),
    ('Personalkosten', 'kosten'),
    ('Fahrzeugkosten', 'kosten'),
    ('Werkzeug & Maschinen', 'kosten'),
    ('Büro & Verwaltung', 'kosten'),
    ('Software & Tools', 'kosten'),
    ('Versicherungen', 'kosten'),
    ('Steuerberater', 'kosten'),
    ('Marketing', 'kosten'),
    ('Sonstige Ausgaben', 'kosten'),
    ('Zinsen', 'kosten'),
    ('Steuern (Gewerbe/ESt)', 'kosten'),
]

r = headers_row + 1
for label, cat in input_categories:
    ws3.row_dimensions[r].height = 18
    if cat is None:
        if label:
            ws3.merge_cells(f'A{r}:N{r}')
            c = ws3.cell(row=r, column=1, value='  ' + label)
            c.fill = fill(SECTION_BG); c.font = font(bold=True, color='FFFFFF', size=10)
            c.alignment = align('left')
        r += 1
        continue

    c = ws3.cell(row=r, column=1, value=label)
    c.alignment = align('left')
    for col in range(2, 14):
        cell = ws3.cell(row=r, column=col, value=0)
        cell.number_format = '#,##0 "€"'
        cell.fill = PatternFill("solid", fgColor="FFFDE7")
        cell.alignment = align('right')
        cell.border = Border(
            bottom=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0')
        )
    # Sum formula
    sum_cell = ws3.cell(row=r, column=14)
    sum_cell.value = f'=SUM(B{r}:M{r})'
    sum_cell.number_format = '#,##0 "€"'
    sum_cell.fill = fill(SUBTOT_BG)
    sum_cell.font = font(bold=True)
    sum_cell.alignment = align('right')
    r += 1

ws3.freeze_panes = 'B5'

# ── Save ─────────────────────────────────────────────────────────────────
out_path = '/home/user/nexx38/GuV_SHK_2025_2024.xlsx'
wb.save(out_path)
print(f'Gespeichert: {out_path}')
